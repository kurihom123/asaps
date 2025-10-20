from django.shortcuts import render, redirect
from ..models import Contribution, Association, ContributionUpload
from ..forms import ExcelUploadForm
from django.db.models import Sum
from collections import defaultdict
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from xhtml2pdf import pisa
import pandas as pd
from datetime import datetime

@login_required
def contribution_list(request):
    # Handle Excel upload
    if request.method == 'POST' and 'excel_file' in request.FILES:
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            return handle_excel_upload(request, form)

    # Handle manual "Add Year" submission
    elif request.method == 'POST' and 'new_year' in request.POST:
        new_year = request.POST.get('new_year').strip()
        if new_year:
            # Check if year already exists in contributions or uploaded records
            existing_years = list(Contribution.objects.values_list('year', flat=True).distinct())
            if new_year not in existing_years:
                # Add a dummy year entry using None values
                messages.success(request, f"Year {new_year} added successfully.")
                request.session.setdefault('manual_years', []).append(new_year)
                request.session.modified = True
            else:
                messages.info(request, f"ℹYear {new_year} already exists.")
        else:
            messages.warning(request, "Please enter a valid year range (e.g. 2025-2026).")

    form = ExcelUploadForm()

    # Fetch contributions and uploads
    contributions = Contribution.objects.select_related('association').all().order_by('year')
    uploads = ContributionUpload.objects.all().order_by('-uploaded_at')

    grouped_contributions = defaultdict(list)
    total_members_by_year = {}
    total_requested_by_year = {}
    total_allocation_by_year = {}
    total_balance_by_year = {}

    # Group contributions by year
    for c in contributions:
        year = c.year
        grouped_contributions[year].append(c)
        total_members_by_year[year] = total_members_by_year.get(year, 0) + c.association.member_number
        total_requested_by_year[year] = total_requested_by_year.get(year, 0) + c.amount_paid
        total_allocation_by_year[year] = total_allocation_by_year.get(year, 0) + c.allocation
        total_balance_by_year[year] = total_balance_by_year.get(year, 0) + c.balance

    # Merge manually added years (from session)
    manual_years = request.session.get('manual_years', [])
    for y in manual_years:
        if y not in grouped_contributions:
            grouped_contributions[y] = []  # empty placeholder

    context = {
        'grouped_contributions': dict(sorted(grouped_contributions.items())),
        'total_members_by_year': total_members_by_year,
        'total_requested_by_year': total_requested_by_year,
        'total_allocation_by_year': total_allocation_by_year,
        'total_balance_by_year': total_balance_by_year,
        'upload_form': form,
        'uploads': uploads,
    }

    return render(request, 'pages/contributions/contribution_list.html', context)


def handle_excel_upload(request, form):
    """Process uploaded Excel file with flexible date handling (no total bills)"""
    excel_file = form.cleaned_data['excel_file']
    year = str(form.cleaned_data['year']).strip()

    try:
        df = pd.read_excel(excel_file)
        df.columns = [str(col).strip() for col in df.columns]

        required_columns = ['Association', 'Members', 'Amount Paid', 'Date Paid']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            messages.error(request, f"Missing columns: {', '.join(missing_columns)}")
            return redirect('contribution_list')

        updated_count, created_count = 0, 0
        errors = []

        for index, row in df.iterrows():
            try:
                assoc_abbr = str(row['Association']).strip()
                members = int(row['Members'])
                amount_paid = float(row['Amount Paid'])

                # --- Handle flexible date format ---
                date_paid = str(row['Date Paid']).strip()
                if pd.isna(date_paid) or date_paid == '':
                    payment_date = '-'  # Blank date
                else:
                    try:
                        payment_date = pd.to_datetime(date_paid, dayfirst=True, errors='coerce')
                        if pd.isna(payment_date):
                            payment_date = '-'
                        else:
                            payment_date = payment_date.date()
                    except Exception:
                        payment_date = '-'

                # --- Find association ---
                try:
                    association = Association.objects.get(abbr=assoc_abbr)
                except Association.DoesNotExist:
                    errors.append(f"Association '{assoc_abbr}' not found.")
                    continue

                # --- Update member count if changed ---
                if association.member_number != members:
                    association.member_number = members
                    association.save()

                # --- Calculate allocation ---
                allocation = members * 500 * 12

                # --- Compute balance directly (no total bills) ---
                balance = allocation - amount_paid

                # --- Save or update contribution ---
                obj, created = Contribution.objects.update_or_create(
                    association=association,
                    year=year,
                    defaults={
                        'allocation': allocation,
                        'amount_paid': amount_paid,
                        'balance': balance,
                        'payment_date': payment_date,
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                continue

        # --- Save upload record ---
        upload = form.save(commit=False)
        upload.uploaded_by = request.user
        upload.save()

        # --- Success messages ---
        messages.success(
            request,
            f"Upload successful ({updated_count} updated, {created_count} new records)."
        )

        if errors:
            for error in errors[:3]:
                messages.warning(request, error)
            if len(errors) > 3:
                messages.warning(request, f"...and {len(errors) - 3} more issues found")

    except Exception as e:
        messages.error(request, f"Upload failed: {str(e)}")

    return redirect('contribution_list')


# Keep your existing PDF and Excel export functions
@login_required
def contributions_pdf(request, year):
    contributions = Contribution.objects.filter(year=year).select_related('association')
    total = contributions.aggregate(
        members=Sum('association__member_number'),
        paid=Sum('amount_paid'),
        allocation=Sum('allocation'),
        balance=Sum('balance'),
    )

    context = {
        'year': year,
        'contributions': contributions,
        'total': total,
    }

    template_path = 'pages/contributions/contribution_pdf.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Contributions_{year}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('PDF generation error', status=500)

    return response

@login_required
def contributions_excel(request, year):
    contributions = Contribution.objects.filter(year=year).select_related('association')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Contributions {year}"

    headers = ['Association', 'Members', 'Amount Paid', 'Allocation', 'Payment Date', 'Balance']
    ws.append(headers)

    for contribution in contributions:
        ws.append([
            contribution.association.abbr,
            contribution.association.member_number,
            contribution.amount_paid,
            contribution.allocation,
            contribution.payment_date.strftime('%Y-%m-%d'),
            contribution.balance,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Contributions_{year}.xlsx'
    wb.save(response)
    return response



@login_required
def my_contributions(request):
    """View contributions for the logged-in user's association"""
    user = request.user

    if user.is_staff:
        return redirect('contribution_list')  # Redirect staff to admin page

    association = getattr(user.user_profile.first(), 'association', None)
    if not association:
        return render(request, 'pages/contributions/my_contributions.html', {
            'error': "No association found for your account."
        })

    contributions = Contribution.objects.filter(association=association).order_by('-year')

    return render(request, 'pages/contributions/my_contributions.html', {
        'association': association,
        'contributions': contributions,
    })


@login_required
def my_arrears(request):
    """View arrears (balances) per year for the user's association"""
    user = request.user

    association = getattr(user.user_profile.first(), 'association', None)
    if not association:
        return render(request, 'pages/contributions/my_arrears.html', {
            'error': "No association found for your account."
        })

    contributions = Contribution.objects.filter(association=association).order_by('-year')
    total_arrears = sum(c.balance for c in contributions)

    return render(request, 'pages/contributions/my_arrears.html', {
        'association': association,
        'contributions': contributions,
        'total_arrears': total_arrears,
    })
