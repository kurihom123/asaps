from django.shortcuts import render, redirect
from ..models import Contribution, Association, ContributionUpload
from ..forms import ExcelUploadForm
from django.db.models import Sum
from collections import defaultdict
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from xhtml2pdf import pisa
import pandas as pd
from datetime import datetime

@login_required
def contribution_list(request):
    if request.method == 'POST' and 'excel_file' in request.FILES:
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            return handle_excel_upload(request, form)
    else:
        form = ExcelUploadForm()

    contributions = Contribution.objects.select_related('association').all().order_by('year')
    uploads = ContributionUpload.objects.all().order_by('-uploaded_at')

    grouped_contributions = defaultdict(list)
    total_members_by_year = {}
    total_requested_by_year = {}
    total_allocation_by_year = {}
    total_balance_by_year = {}

    for contribution in contributions:
        year = contribution.year
        grouped_contributions[year].append(contribution)

        total_members_by_year[year] = total_members_by_year.get(year, 0) + contribution.association.member_number
        total_requested_by_year[year] = total_requested_by_year.get(year, 0) + contribution.amount_paid
        total_allocation_by_year[year] = total_allocation_by_year.get(year, 0) + contribution.allocation
        total_balance_by_year[year] = total_balance_by_year.get(year, 0) + contribution.balance

    context = {
        'grouped_contributions': dict(grouped_contributions),
        'total_members_by_year': total_members_by_year,
        'total_requested_by_year': total_requested_by_year,
        'total_allocation_by_year': total_allocation_by_year,
        'total_balance_by_year': total_balance_by_year,
        'upload_form': form,
        'uploads': uploads,
    }
    return render(request, 'pages/contributions/contribution_list.html', context)


def handle_excel_upload(request, form):
    """Process the uploaded Excel file and create/update contributions"""
    excel_file = form.cleaned_data['excel_file']
    year = form.cleaned_data['year']
    
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file)
        
        # Validate required columns
        required_columns = ['Association', 'Members', 'Amount Paid']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            messages.error(request, f"Missing columns in Excel file: {', '.join(missing_columns)}")
            return redirect('contribution_list')
        
        success_count = 0
        update_count = 0
        error_messages = []
        
        for index, row in df.iterrows():
            try:
                association_abbr = str(row['Association']).strip()
                members = int(row['Members'])
                new_amount_paid = float(row['Amount Paid'])
                
                # Handle Payment Date
                payment_date = row.get('Payment Date')
                if pd.isna(payment_date) or payment_date == '':
                    payment_date = datetime.now().date()
                elif isinstance(payment_date, str):
                    payment_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
                
                # Get association
                try:
                    association = Association.objects.get(abbr=association_abbr)
                except Association.DoesNotExist:
                    error_messages.append(f"Association '{association_abbr}' not found")
                    continue
                
                # Calculate allocation (500 * members * 12)
                allocation = members * 500 * 12
                
                # Check if contribution already exists
                existing_contribution = Contribution.objects.filter(
                    association=association, 
                    year=year
                ).first()
                
                if existing_contribution:
                    # OPTION 1: REPLACE the amount (use this if you want to overwrite)
                    total_amount_paid = new_amount_paid
                    
                    # OPTION 2: ADD to existing amount (uncomment if you want cumulative)
                    # total_amount_paid = existing_contribution.amount_paid + new_amount_paid
                    
                    # Update the contribution
                    existing_contribution.amount_paid = total_amount_paid
                    existing_contribution.allocation = allocation
                    existing_contribution.balance = allocation - total_amount_paid
                    existing_contribution.payment_date = payment_date
                    
                    # Update member number if changed
                    if association.member_number != members:
                        association.member_number = members
                        association.save()
                    
                    existing_contribution.save()
                    update_count += 1
                    messages.info(request, f"Updated {association_abbr}: Amount paid = {total_amount_paid:,}, Balance = {allocation - total_amount_paid:,}")
                    
                else:
                    # Create new contribution
                    if association.member_number != members:
                        association.member_number = members
                        association.save()
                    
                    contribution = Contribution.objects.create(
                        association=association,
                        year=year,
                        allocation=allocation,
                        amount_paid=new_amount_paid,
                        payment_date=payment_date,
                        balance=allocation - new_amount_paid
                    )
                    success_count += 1
                    messages.info(request, f"Added {association_abbr}: Amount paid = {new_amount_paid:,}, Balance = {allocation - new_amount_paid:,}")
                
            except ValueError as e:
                error_messages.append(f"Row {index + 2}: Invalid number format - {str(e)}")
                continue
            except Exception as e:
                error_messages.append(f"Row {index + 2}: {str(e)}")
                continue
        
        # Save upload record
        upload = form.save(commit=False)
        upload.uploaded_by = request.user
        upload.save()
        
        # Show summary messages
        if success_count > 0:
            messages.success(request, f"Added {success_count} new contributions for {year}")
        if update_count > 0:
            messages.success(request, f"Updated {update_count} existing contributions for {year}")
        if not success_count and not update_count and not error_messages:
            messages.warning(request, "No data was processed from the Excel file")
        if error_messages:
            for error in error_messages[:5]:
                messages.warning(request, error)
            if len(error_messages) > 5:
                messages.warning(request, f"... and {len(error_messages) - 5} more errors")
        
    except Exception as e:
        messages.error(request, f"Error processing Excel file: {str(e)}")
    
    return redirect('contribution_list')

# Keep your existing PDF and Excel export functions
@login_required
def contributions_pdf(request, year):
    contributions = Contribution.objects.filter(year=year).select_related('association')
    total = contributions.aggregate(
        members=Sum('association__member_number'),
        paid=Sum('amount_paid'),
        allocation=Sum('allocation'),
        balance=Sum('balance'),
    )

    context = {
        'year': year,
        'contributions': contributions,
        'total': total,
    }

    template_path = 'pages/contributions/contribution_pdf.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Contributions_{year}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('PDF generation error', status=500)

    return response

@login_required
def contributions_excel(request, year):
    contributions = Contribution.objects.filter(year=year).select_related('association')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Contributions {year}"

    headers = ['Association', 'Members', 'Amount Paid', 'Allocation', 'Payment Date', 'Balance']
    ws.append(headers)

    for contribution in contributions:
        ws.append([
            contribution.association.abbr,
            contribution.association.member_number,
            contribution.amount_paid,
            contribution.allocation,
            contribution.payment_date.strftime('%Y-%m-%d'),
            contribution.balance,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Contributions_{year}.xlsx'
    wb.save(response)
    return response

# Remove the add_contribution_for_year function